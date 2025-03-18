import pandas as pd
from functions import get_sonar_data, get_snyk_data, get_checkmarx_data
from openpyxl import load_workbook

file_path = "/Users/seanbowen/Downloads/test.xlsx"
df = pd.read_excel(file_path, sheet_name="Sheet1")  # Adjust sheet name
repo_urls = df["Repository URL"].tolist()

results = []
for url in repo_urls:
    repo_name = url.split("/")[-1]  # Extract repo name from URL
    sonar_data = get_sonar_data(repo_name)
    # snyk_data = get_snyk_data(repo_name)
    # checkmarx_data = get_checkmarx_data(repo_name)

    results.append({
        "Repository URL": url,
        "Code Coverage": sonar_data["code_coverage"],
        "Code Smells": sonar_data["code_smells"],
        "Security Hotspots": sonar_data["security_hotspots"],
        # "Snyk Critical Vulns": snyk_data["critical_vulns"],
        # "Snyk High Vulns": snyk_data["high_vulns"],
        # "Checkmarx Critical": checkmarx_data["checkmarx_critical"],
        # "Checkmarx High": checkmarx_data["checkmarx_high"]
    })

# Convert results to DataFrame and update Excel
new_df = pd.DataFrame(results)
wb = load_workbook(file_path)
ws = wb["Sheet1"]
for r_idx, row in enumerate(new_df.itertuples(index=False), 2):  # Start from row 2
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)
wb.save(file_path)